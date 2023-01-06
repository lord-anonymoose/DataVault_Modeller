import sys
import os
import openpyxl
from PySide6.QtGui import QGuiApplication
from PySide6.QtQml import QQmlApplicationEngine
from PySide6.QtCore import QObject, Signal, Slot

class MainBackend(QObject):
    path1 = Signal(str, arguments=['write'])

    def __init__(self):
        QObject.__init__(self)

    # Logical Data Model Check
    @Slot('QString', 'QString')
    def checkModel(self, pathLMD, pathSave):
        # Uploading Logical Data Model Check
        wb = openpyxl.load_workbook(filename = pathLMD)
        sheet = wb['Модель']
        errors = []
        # Data Vault Standards
        stereotypes = ['HUB', 'SAT', 'LNK']
        data_types = [
            'bigint',
            'bigserial',
            'boolean',
            'box',
            'char',
            'varchar',
            'cidr',
            'date',
            'cidr',
            'real',
            'polygon',
            'real',
            'serial',
            'smallint',
            'text',
            'timestamp',
            'uuid'
        ]
        goodStructure = True
        # Unknown symbols checker
        def match (text, alphabet=set('abcdefghijklmnopqrstuvwxyz_0123456789')):
            return not alphabet.isdisjoint(text.lower())
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 1):
            for cell in row:
               if not match(cell.value):
                   errors.append("Схема {} имеет некорректное название".format(cell.value))
        # Column name check
        def checkColumn (number, name):
            value = sheet.cell(row = 1, column = number).value
            return value == name
        # Full structure check
        goodStructure = checkColumn(1, "table_schema") and checkColumn(2, "table_type") and checkColumn(3, "table_name") and checkColumn(4, "table_desc") and checkColumn(5, "field_name") and checkColumn(6, "field_type") and checkColumn(7, "null / not null") and checkColumn(8, "field_desc") and checkColumn(9, "distribution / partition")
        # Check for unknown stereotypes
        for row in sheet.iter_rows(min_row = 2, min_col = 2, max_col = 2):
            for cell in row:
                if cell.value not in stereotypes:
                    errors.append("Некорректное значение стереотипа таблицы: {}".format(cell.value))
        # Check for unknown symbols in table names
        for row in sheet.iter_rows(min_row = 2, min_col = 3, max_col = 3):
            for cell in row:
                if not match(cell.value):
                    errors.append("Таблица {} имеет некорректное название".format(cell.value))
        # Check for empty table descriptions
        for row in sheet.iter_rows(min_row = 2, min_col = 4, max_col = 4):
            for cell in row:
                if not cell.value:
                    errors.append("Встречаются незаполненные описания таблиц")
        # Check for unknown symbols in column names
        for row in sheet.iter_rows(min_row = 2, min_col = 5, max_col= 5):
            for cell in row:
                if not match(cell.value):
                    errors.append("Таблица {} имеет некорректное название".format(cell.value))
        # Check for unknown data types
        for row in sheet.iter_rows(min_row = 2, min_col = 6, max_col = 6):
            for cell in row:
                if cell.value not in data_types:
                    errors.append("Несуществующий тип данных: {}".format(cell.value))
        # Check for unknown values in "null / not null" column
        for row in sheet.iter_rows(min_row = 2, min_col = 7, max_col = 7):
            for cell in row:
                if cell.value not in ['null', 'not null']:
                    errors.append("Некорректное значение в столбце null / not null: {}".format(cell.value))
        # Check for empty column descriptions
        for row in sheet.iter_rows(min_row = 2, min_col = 8, max_col = 8):
            for cell in row:
                if not cell.value:
                    errors.append("Встречаются незаполненные описания столбцов")
        # Check for column duplicates
        columns = []
        unique_columns = []
        for row in sheet.iter_rows(min_row = 2, min_col = 3, max_col = 3):
            for cell in row:
                other_cell = sheet.cell(row = cell.row, column = 5)
                columns.append(cell.value + "." + other_cell.value)
        for i in columns:
            if i not in unique_columns:
                unique_columns.append(i)
            else:
                errors.append("Дублируется поле: {}".format(i))
        # Error report output
        os.chdir(pathSave)
        with open('error_report.txt', 'w') as f:
            if goodStructure:
                for error in errors:
                    f.write(f"{error}\n")
                if not errors:
                    f.write("Модель успешно прошла проверки")
            else:
                f.write("Некорректная структура файла модели")

    # DDL-structures Generator
    @Slot('QString', 'QString')
    def generateDatabase(self, pathLMD, pathSave):
        ## Подгружается файл с моделью
        wb = openpyxl.load_workbook(filename = pathLMD)
        sheet = wb['Модель']
        # Creating a list of unique table names
        tables = []
        unique_tables = []
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 1):
            for cell in row:
                other_cell = sheet.cell(row = cell.row, column = 3)
                tables.append(cell.value + "." + other_cell.value)
        for i in tables:
            if i not in unique_tables:
                unique_tables.append(i)
        # Creating a dictionary for storing scripts
        scripts = {}
        for i in unique_tables:
            scripts[i] = "create table " + i + " ("
        # Filling in scripts dictionary
        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_col = 1):
            for cell in row:
                table_schema = sheet.cell(row = cell.row, column = 1)
                table_name = sheet.cell(row = cell.row, column = 3)
                column = sheet.cell(row = cell.row, column = 5)
                data_type = sheet.cell(row = cell.row, column = 6)
                nullable = sheet.cell(row = cell.row, column = 7)
                scripts[table_schema.value + "." + table_name.value] += ("\n\t" + column.value + " " + data_type.value + " " + nullable.value + ",")
        # DDL-scripts output
        os.chdir(pathSave)
        with open('scripts.txt', 'w') as f:
            for table in scripts:
                f.write("\n")
                f.write("\n")
                for script in scripts[table]:
                    f.write(f"{script}")
                f.write(");")
        f.close()
        f = open('scripts.txt', 'r')
        data = f.read()
        data = data.replace(",)", "\n)")
        f.close()
        f = open('scripts.txt', 'w')
        f.write(data)
        f.close()

if __name__ == "__main__":
    app = QGuiApplication(sys.argv)
    engine = QQmlApplicationEngine()
    backend = MainBackend()
    engine.rootContext().setContextProperty("backend", backend)
    engine.load("main.qml")
    engine.quit.connect(app.quit)
    sys.exit(app.exec_())
